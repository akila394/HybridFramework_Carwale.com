import openpyxl

def getRawCount(path,sheet):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return  sheet.max_row

def getRColumnCount(path,sheet):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return  sheet.max_column

def getCellData(path, sheet, rownumber,columnnumber):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    return sheet.cell(row = rownumber,column = columnnumber).value

def setCellData(path, sheet, rownumber, columnnumber, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    sheet.cell(row = rownumber,column = columnnumber).value = data
    workbook.save("..//excelfiles//testdata.xlsx")




path = "..//excelfiles//testdata.xlsx"
sheet = "logintest"

print("row count : ",getRawCount(path,sheet))
print("column count : ", getRColumnCount(path,sheet))
print(getCellData(path,sheet,3,1))
setCellData(path,sheet,4,1,"ruwanipd@gmail.com")